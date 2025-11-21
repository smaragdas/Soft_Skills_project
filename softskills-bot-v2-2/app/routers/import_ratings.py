from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlmodel import Session, select
from typing import Optional
from openpyxl import load_workbook
from app.core.db import get_session
from app.models.db_models import Interaction, HumanRating

router = APIRouter(prefix="/import", tags=["import"])

@router.post("/human-xlsx")
async def import_human_xlsx(
    file: UploadFile = File(...),
    rater_id: Optional[str] = Form(None),
    session: Session = Depends(get_session)
):
    content = await file.read()
    wb = load_workbook(filename=None, data=content)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
    # map header names to indices
    col = {name: idx for idx, name in enumerate(headers)}
    required = ["answer_id", "score"]
    for r in required:
        if r not in col:
            raise HTTPException(status_code=400, detail=f"Missing required column: {r}")

    inserted = 0
    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        ans_id = row[col["answer_id"]]
        if not ans_id:
            continue
        score_val = row[col["score"]]
        if score_val is None:
            skipped += 1
            continue
        try:
            score = float(score_val)
        except:
            skipped += 1
            continue
        rid = rater_id or (row[col["rater_id"]] if "rater_id" in col else None)
        if not rid:
            rid = "r0"

        notes = row[col["notes"]] if "notes" in col else None

        inter = session.exec(select(Interaction).where(Interaction.answer_id == ans_id)).first()
        if not inter:
            skipped += 1
            continue

        existing = session.exec(select(HumanRating).where((HumanRating.answer_id == ans_id) & (HumanRating.rater_id == rid))).first()
        if existing:
            existing.score = score
            existing.notes = notes
            updated += 1
        else:
            hr = HumanRating(answer_id=ans_id, rater_id=rid, score=score, notes=notes)
            session.add(hr)
            inserted += 1

    session.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}
